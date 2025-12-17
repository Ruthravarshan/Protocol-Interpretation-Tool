# annotation_extractor.py
import json
from typing import List, Set

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def extract_annotations_from_excel(file_path: str) -> str:
    """
    Extract annotations from Excel cells (superscript text).
    
    Rules:
    - Find all superscript text in cells
    - Split by commas and trim whitespace
    - Keep only non-empty values
    - Ignore annotations longer than 3 characters
    - Sort numerically if all numeric, else lexicographically
    - Remove duplicates while preserving sorted order
    - Return as JSON array of strings
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        JSON array string (e.g., '["1","3","5"]')
    """
    if not HAS_OPENPYXL:
        return json.dumps([])
    
    annotations = set()
    
    try:
        # Load workbook with data_only=False to get formatting info
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Iterate through all cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                    
                cell_text = str(cell.value)
                
                # Check if cell has superscript formatting
                if cell.font and cell.font.superscript:
                    # Extract superscript annotations
                    _extract_and_add_annotations(cell_text, annotations)
                else:
                    # Try to extract superscript from mixed text
                    # Look for patterns that might indicate superscript
                    extracted = _extract_superscript_annotations(cell_text)
                    if extracted:
                        for ann in extracted:
                            _add_annotation(ann, annotations)
        
        wb.close()
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return json.dumps([])
    
    # Process annotations: split, trim, filter, sort, deduplicate
    final_annotations = _process_annotations(annotations)
    
    return json.dumps(final_annotations)


def _extract_superscript_annotations(text: str) -> List[str]:
    """
    Try to detect superscript patterns in text.
    Common patterns: text¹, text², text¹'², etc.
    """
    # Unicode superscript characters
    superscript_map = {
        '⁰': '0', '¹': '1', '²': '2', '³': '3', '⁴': '4',
        '⁵': '5', '⁶': '6', '⁷': '7', '⁸': '8', '⁹': '9',
        '⁺': '+', '⁻': '-', '⁽': '(', '⁾': ')'
    }
    
    annotations = []
    current_superscript = ""
    
    for char in text:
        if char in superscript_map:
            current_superscript += superscript_map[char]
        else:
            if current_superscript:
                annotations.append(current_superscript)
                current_superscript = ""
    
    if current_superscript:
        annotations.append(current_superscript)
    
    return annotations


def _extract_and_add_annotations(text: str, annotations: Set[str]):
    """
    Extract annotations from cell text by splitting on commas.
    """
    parts = text.split(',')
    for part in parts:
        _add_annotation(part, annotations)


def _add_annotation(text: str, annotations: Set[str]):
    """
    Process a single annotation: trim, filter by length, add to set.
    """
    trimmed = text.strip()
    if trimmed and len(trimmed) <= 3:
        annotations.add(trimmed)


def _process_annotations(annotations: Set[str]) -> List[str]:
    """
    Process annotations: split further by commas if needed, filter, sort, deduplicate.
    
    Rules:
    - Split by commas and trim
    - Keep only non-empty values <= 3 chars
    - Sort numerically if all numeric, else lexicographically
    - Remove duplicates
    """
    if not annotations:
        return []
    
    # Process: split by commas where needed
    all_parts = set()
    for ann in annotations:
        if ',' in ann:
            parts = [p.strip() for p in ann.split(',')]
            for p in parts:
                if p and len(p) <= 3:
                    all_parts.add(p)
        else:
            if ann and len(ann) <= 3:
                all_parts.add(ann)
    
    # Check if all are numeric
    all_numeric = all(p.isdigit() for p in all_parts)
    
    # Sort
    if all_numeric:
        # Sort numerically
        sorted_annotations = sorted(all_parts, key=lambda x: int(x))
    else:
        # Sort lexicographically
        sorted_annotations = sorted(all_parts)
    
    return sorted_annotations


def extract_annotations_from_dataframe(df) -> str:
    """
    Extract annotations from pandas DataFrame (used when Excel is not available).
    Looks for superscript Unicode characters or special patterns.
    
    Args:
        df: pandas DataFrame
        
    Returns:
        JSON array string
    """
    try:
        import pandas as pd
    except ImportError:
        return json.dumps([])
    
    annotations = set()
    
    # Iterate through all cells
    try:
        for row_idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                    continue
                
                cell_text = str(cell)
                
                # Try to extract superscript annotations
                extracted = _extract_superscript_annotations(cell_text)
                if extracted:
                    for ann in extracted:
                        _add_annotation(ann, annotations)
    except Exception as e:
        print(f"Error iterating dataframe: {e}")
    
    # Process annotations
    final_annotations = _process_annotations(annotations)
    
    return json.dumps(final_annotations)
